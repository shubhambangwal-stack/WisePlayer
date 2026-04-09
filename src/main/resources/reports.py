import subprocess
import os
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

REPORT_DIR = "reports"


def run_git_command(cmd):
    try:
        result = subprocess.check_output(cmd, shell=True, text=True).strip()
        return result
    except subprocess.CalledProcessError:
        return ""


def get_today_git_commits():
    today = date.today().strftime("%Y-%m-%d")

    cmd = f'git log --since="{today} 00:00" --until="{today} 23:59" --pretty=format:"%h|%an|%ad|%s" --date=short'
    output = run_git_command(cmd)

    commits = []
    if output:
        for line in output.split("\n"):
            parts = line.split("|", 3)
            if len(parts) == 4:
                commits.append({
                    "hash": parts[0],
                    "author": parts[1],
                    "date": parts[2],
                    "message": parts[3]
                })
    return commits


def ask_manual_tasks():
    tasks = []
    print("\nEnter today's tasks manually (type 'done' to finish):\n")

    while True:
        title = input("Task Title (or 'done'): ").strip()
        if title.lower() == "done":
            break

        desc = input("Description: ").strip()
        start = input("Start Time (HH:MM): ").strip()
        end = input("End Time (HH:MM): ").strip()
        remarks = input("Remarks/Outcome: ").strip()

        tasks.append({
            "title": title,
            "description": desc,
            "start": start,
            "end": end,
            "remarks": remarks
        })
        print("-" * 40)

    return tasks


def calculate_time_diff(start, end):
    try:
        fmt = "%H:%M"
        start_dt = datetime.strptime(start, fmt)
        end_dt = datetime.strptime(end, fmt)

        diff = end_dt - start_dt
        total_minutes = diff.total_seconds() / 60

        if total_minutes < 0:
            return "Invalid"

        hours = int(total_minutes // 60)
        minutes = int(total_minutes % 60)

        return f"{hours}h {minutes}m"
    except:
        return "Invalid"


def auto_adjust_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 3


def generate_excel(tasks, commits):
    if not os.path.exists(REPORT_DIR):
        os.makedirs(REPORT_DIR)

    today_str = date.today().strftime("%Y-%m-%d")
    filename = f"{REPORT_DIR}/Daily_Report_{today_str}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    # Header
    ws["A1"] = "Daily Work Report"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Date: {today_str}"
    ws["A2"].font = Font(bold=True)

    row = 4

    # Manual Tasks Section
    ws[f"A{row}"] = "Manual Tasks"
    ws[f"A{row}"].font = Font(size=14, bold=True)
    row += 1

    headers = ["Task Title", "Description", "Start Time", "End Time", "Time Spent", "Remarks/Outcome"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row += 1

    total_minutes = 0

    for task in tasks:
        spent = calculate_time_diff(task["start"], task["end"])

        ws.cell(row=row, column=1, value=task["title"])
        ws.cell(row=row, column=2, value=task["description"])
        ws.cell(row=row, column=3, value=task["start"])
        ws.cell(row=row, column=4, value=task["end"])
        ws.cell(row=row, column=5, value=spent)
        ws.cell(row=row, column=6, value=task["remarks"])

        # total time calculation
        try:
            fmt = "%H:%M"
            s = datetime.strptime(task["start"], fmt)
            e = datetime.strptime(task["end"], fmt)
            mins = (e - s).total_seconds() / 60
            if mins > 0:
                total_minutes += mins
        except:
            pass

        row += 1

    row += 2

    # Total time summary
    ws[f"A{row}"] = "Total Manual Work Time:"
    ws[f"A{row}"].font = Font(bold=True)

    hours = int(total_minutes // 60)
    minutes = int(total_minutes % 60)

    ws[f"B{row}"] = f"{hours}h {minutes}m"
    ws[f"B{row}"].font = Font(bold=True)

    row += 3

    # Git Commits Section
    ws[f"A{row}"] = "Git Commits (Today)"
    ws[f"A{row}"].font = Font(size=14, bold=True)
    row += 1

    commit_headers = ["Commit Hash", "Author", "Date", "Message"]
    for col, header in enumerate(commit_headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row += 1

    if commits:
        for c in commits:
            ws.cell(row=row, column=1, value=c["hash"])
            ws.cell(row=row, column=2, value=c["author"])
            ws.cell(row=row, column=3, value=c["date"])
            ws.cell(row=row, column=4, value=c["message"])
            row += 1
    else:
        ws.cell(row=row, column=1, value="No commits found today.")

    auto_adjust_columns(ws)

    wb.save(filename)
    print(f"\n✅ Report generated: {filename}")


if __name__ == "__main__":
    print("==== Daily Report Generator ====")

    print("\nFetching today's git commits...")
    commits = get_today_git_commits()
    print(f"Found {len(commits)} commits.")

    tasks = ask_manual_tasks()

    generate_excel(tasks, commits)